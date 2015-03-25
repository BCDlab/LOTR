

from openpyxl import Workbook
from openpyxl import load_workbook


wb = load_workbook('crdi labeling_with charisse_bookIP.xlsx')
ws = wb.active


for rowCount in range(2, 241):
	if ((rowCount - 2) % 8) == 0:
		cellValue = ws['B{0}'.format(rowCount)].value

		name = ''
		
		if cellValue == 'book_left' or cellValue == 'book_right':
			name = 'book'
		elif cellValue == 'bottle_left' or cellValue == 'bottle_right':
			name = 'bottle'

		print name

		for i in range(rowCount, rowCount + 8):
			ws['Y{0}'.format(i)] = name


wb.save('output.xlsx')
