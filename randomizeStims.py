
#################################################################

######################
###  Description
######################

# BCD Lab 2015, Jim Calabro
# For assistance email jcalabro@umass.edu

# Program designed to generate a pseudo-randomly ordered
# excel sheet of stimuli pairings

# SPECIFICATIONS:
#   Objects:
#      -No two "species" should be paired together
#      -Don't want colors to be the same
#      -Once pairing has happened, it shouldn't happen again
#   Faces:
#      -One trained, one untrained
#      -Don't have consecutive faces of the same race
#      -Have each race look in each direction
#      -Don't look consecutively in the same direction more than once

#################################################################

######################
###  Configuration
######################


# import statements
from openpyxl import load_workbook
from random import randint

# where to read and write the excel files
read_location = 'C:/Users/Jim/Code/BCD/LOTR/Stimuli/'
write_location = read_location

# name of the workbook to be read
bird_workbook_name = 'ObjectsUsed_bird_color.xlsx'

# the number of birds in 
number_of_birds = 24

bird_workbook = load_workbook('{0}{1}'.format(read_location, bird_workbook_name))
bird_ws = bird_workbook.active

output_bird_workbook = load_workbook('{0}birdOutputTemplate.xlsx'.format(read_location))

#################################################################

######################
###  Execution
######################


bird_list = []
pair_list = []

cell_range = bird_ws['A2:E{0}'.format(number_of_birds + 1)]
for bird in cell_range:
    bird_list.append(bird)

# REMOVE LATER
# for i in range(0,24):
    # print bird_list[i]

print



# hard-code the first random stimuli, then iterate to fill in the list

pair_count = 0

random_left_bird_1 = bird_list[randint(0, len(bird_list) - 1)]
random_right_bird_1 = bird_list[randint(0, len(bird_list) - 1)]

# make sure a and b are different species and different colors
while random_left_bird_1[4].value == random_right_bird_1[4].value or random_left_bird_1[2].value == random_right_bird_1[2].value:
    random_right_bird_1 = bird_list[randint(0, len(bird_list) - 1)]

pair_list.append((random_left_bird_1, random_right_bird_1))

# print pair_list

pair_count += 1

# iterate to get the rest of the stimuli

while len(pair_list) < number_of_birds:
	random_a_bird = bird_list[randint(0, len(bird_list) - 1)]
	random_b_bird = bird_list[randint(0, len(bird_list) - 1)]

	current_pair = [random_a_bird, random_b_bird]

	if pair_list[pair_count - 1][0] != random_a_bird[0] and random_a_bird != random_b_bird and current_pair not in pair_list: 
	    pair_list.append((current_pair[0], current_pair[1]))
	# else:
	    # print 'Match invalid; repeating random pairing


#      -No two "species" should be paired together - Done
#      -Don't want colors to be the same
#      -Once pairing has happened, it shouldn't happen again

#################################################################

######################
###  Save Data
######################

# write pairings to new file using a template

ws = output_bird_workbook.active

# start at the second row
row_count = 2

while len(pair_list) > 1:
	cell_range_left  = ws['A{0}:E{0}'.format(row_count)]
	cell_range_right = ws['F{0}:J{0}'.format(row_count)]

	current_pair = pair_list.pop(0)

	column_count = 0

	for row in cell_range_left:
		for cell in row:
			# print cell.value
			cell.value = current_pair[0][column_count].value
			column_count += 1

	column_count = 0

	for row in cell_range_right:
		for cell in row:
			# print cell.value
			cell.value = current_pair[1][column_count].value
			column_count += 1

	row_count += 1

output_bird_workbook.save('{0}output.xlsx'.format(write_location))

# finished, exit successfully

print 'Finished; exit code 0'
